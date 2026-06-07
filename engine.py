"""
engine.py  –  Rodeo Checker core matching engine
Produces the Fines_Card_Verification report.

Sheets:
  ⚑ Summary           – stat boxes + full flagged persons table
  ✅ Entries With Cards  – every alpha-sheet name that has a card number
  ❌ Entries Without Cards – every alpha-sheet name with no card number
  💰 Fine Totals       – every fine from the suspended list, per person
"""

import re
import unicodedata
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from collections import OrderedDict
import os


# ── Name normalisation ────────────────────────────────────────────────────────

def _norm(text) -> str:
    if not isinstance(text, str) or not text.strip():
        return ""
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9 ]", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def make_key(last, first) -> str:
    return _norm(str(last)) + "|||" + _norm(str(first))


# ── File parsers ──────────────────────────────────────────────────────────────

def load_alpha_sheet(path: str) -> pd.DataFrame:
    import csv as _csv
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        # Use Python csv reader so variable-length TeamRoping rows (which carry
        # heeler name in cols 7-8 beyond the 7-col header) are not dropped.
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = _csv.reader(fh)
            raw_rows = [r for r in reader]
        if not raw_rows:
            return pd.DataFrame()
        header = raw_rows[0]
        records = []
        for row in raw_rows[1:]:
            # Skip blank rows and rows containing only NUL/whitespace bytes
            printable = [c for c in row if c.strip() and all(ord(ch) >= 32 for ch in c)]
            if not printable:
                continue
            rec = {header[i]: row[i] for i in range(min(len(header), len(row)))}
            rec["_pos7"] = row[7] if len(row) > 7 else ""
            rec["_pos8"] = row[8] if len(row) > 8 else ""
            records.append(rec)
        df = pd.DataFrame(records)
    else:
        df = pd.read_excel(path)
        df["_pos7"] = ""
        df["_pos8"] = ""

    col_map = {}
    for c in df.columns:
        n = _norm(str(c))
        if "RIDER LAST" in n or n == "LAST NAME":
            col_map["last"] = c
        elif "RIDER FIRST" in n or n == "FIRST NAME":
            col_map["first"] = c
        elif "CLASS" in n:
            col_map["class"] = c
        elif "ENTRY TIME" in n or "ENTRY" in n:
            col_map["time"] = c

    rows = []
    for _, r in df.iterrows():
        last  = _norm(str(r.get(col_map.get("last",  "Rider Last Name"),  "")))
        first = _norm(str(r.get(col_map.get("first", "Rider First Name"), "")))
        if not last:
            continue
        cls        = str(r.get(col_map.get("class", "Class Name"), "")).strip()
        entry_time = str(r.get(col_map.get("time",  "Entry Time"),  "")).strip()
        is_tr      = "teamroping" in cls.lower().replace(" ", "")

        rows.append({
            "key":        make_key(last, first),
            "last":       last,
            "first":      first,
            "class_name": cls,
            "entry_time": entry_time,
            "role":       "Header" if is_tr else "",
        })

        # For TeamRoping rows, also add the heeler from the extra positional columns
        if is_tr:
            h_last  = _norm(str(r.get("_pos7", "")))
            h_first = _norm(str(r.get("_pos8", "")))
            if h_last and h_last not in ("NAN", ""):
                rows.append({
                    "key":        make_key(h_last, h_first),
                    "last":       h_last,
                    "first":      h_first,
                    "class_name": cls,
                    "entry_time": entry_time,
                    "role":       "Heeler",
                })

    return pd.DataFrame(rows)


def load_card_numbers(path: str) -> dict:
    """
    Two member blocks side by side:
      Left:  col0=card#  col1=last  col2=first  col3=city  col4=state  col5=events
      Right: col7=card#  col8=last  col9=first  col10=city col11=state col12=events
    Data starts row index 2.
    Returns dict: key -> list of {card_number, last, first, city, state, events}
    """
    df = pd.read_excel(path, header=None)
    lookup = {}

    def _add(card, last, first, city, state, events):
        l = _norm(str(last))
        f = _norm(str(first))
        if not l or l in ("NAN", "UPRA MEMBERS 2026", "NEW"):
            return
        card_str = str(card).strip().split(".")[0]
        if card_str in ("nan", "NAN", ""):
            return
        k = make_key(l, f)
        entry = {
            "card_number": card_str,
            "last": l, "first": f,
            "city":   str(city).strip()   if str(city)   not in ("nan","NaN") else "",
            "state":  str(state).strip()  if str(state)  not in ("nan","NaN") else "",
            "events": str(events).strip() if str(events) not in ("nan","NaN") else "",
        }
        if entry not in lookup.get(k, []):
            lookup.setdefault(k, []).append(entry)

    for _, r in df.iloc[2:].iterrows():
        vals = r.tolist()
        _add(vals[0], vals[1], vals[2],
             vals[3] if len(vals) > 3 else "",
             vals[4] if len(vals) > 4 else "",
             vals[5] if len(vals) > 5 else "")
        if len(vals) > 7 and pd.notna(vals[7]):
            _add(vals[7],  vals[8],  vals[9],
                 vals[10] if len(vals) > 10 else "",
                 vals[11] if len(vals) > 11 else "",
                 vals[12] if len(vals) > 12 else "")
    return lookup


def load_suspended(path: str) -> dict:
    """
    Row 3 = header (LAST NAME / FIRST NAME / OFFENSE / AMOUNT / EVENT)
    Row 4+ = data.
    Returns dict: key -> list of {last, first, offense, amount, event}
    """
    df = pd.read_excel(path, header=None)
    header_idx = 3
    for i, r in df.iterrows():
        if "LAST NAME" in [_norm(str(v)) for v in r.values]:
            header_idx = i
            break

    lookup = {}
    for _, r in df.iloc[header_idx + 1:].iterrows():
        vals = r.tolist()
        last  = _norm(str(vals[0])) if len(vals) > 0 else ""
        first = _norm(str(vals[1])) if len(vals) > 1 else ""
        if not last or last == "NAN":
            continue
        offense = str(vals[2]).strip() if len(vals) > 2 else ""
        try:
            amount = float(str(vals[3]).replace("$","").replace(",","") or 0)
        except (ValueError, TypeError):
            amount = 0.0
        event = str(vals[4]).strip() if len(vals) > 4 else ""
        k = make_key(last, first)
        lookup.setdefault(k, []).append({
            "last": last, "first": first,
            "offense": offense, "amount": amount, "event": event,
        })
    return lookup


# ── Matching ──────────────────────────────────────────────────────────────────

def run_match(alpha_path: str, card_path: str, susp_path: str):
    alpha     = load_alpha_sheet(alpha_path)
    cards     = load_card_numbers(card_path)
    suspended = load_suspended(susp_path)

    # ── Build per-person records for everyone on the alpha sheet ──────────────
    # Key: (last, first) -> record
    person_map = OrderedDict()

    for _, rider in alpha.iterrows():
        k     = rider["key"]
        last  = rider["last"]
        first = rider["first"]
        cls   = rider["class_name"]
        role  = rider.get("role", "")

        if k not in person_map:
            person_map[k] = {
                "last":         last,
                "first":        first,
                "classes":      set(),
                "card_entries": [],
                "susp_entries": [],
                "on_card_list": False,
                "on_susp_list": False,
            }

        # Store class with role suffix for TeamRoping participants
        cls_label = f"{cls} ({role})" if role else cls
        person_map[k]["classes"].add(cls_label)

        for ch in cards.get(k, []):
            person_map[k]["on_card_list"] = True
            entry = {k2: ch[k2] for k2 in ("card_number","events","city","state")}
            if entry not in person_map[k]["card_entries"]:
                person_map[k]["card_entries"].append(entry)

        for sh in suspended.get(k, []):
            person_map[k]["on_susp_list"] = True
            person_map[k]["susp_entries"].append({
                "offense": sh["offense"],
                "amount":  sh["amount"],
                "event":   sh["event"],
            })

    # ── Split into with-card and without-card lists ───────────────────────────
    with_card    = {k: p for k, p in person_map.items() if p["on_card_list"]}
    without_card = {k: p for k, p in person_map.items() if not p["on_card_list"]}

    # ── Fine totals: ALL suspended matches from alpha sheet ───────────────────
    fines_persons = {k: p for k, p in person_map.items() if p["on_susp_list"]}

    # ── Summary flagged = anyone on either reference list ────────────────────
    flagged = {k: p for k, p in person_map.items()
               if p["on_card_list"] or p["on_susp_list"]}

    stats = {
        "total_entrants":   len(alpha),
        "total_unique":     len(set(person_map.keys())),
        "with_card":        len(with_card),
        "without_card":     len(without_card),
        "susp_matches":     len(fines_persons),
        "total_owed":       sum(
                                sum(e["amount"] for e in p["susp_entries"])
                                for p in fines_persons.values()
                            ),
        "total_violations": sum(len(p["susp_entries"]) for p in fines_persons.values()),
        "flagged_names":    len(flagged),
    }

    return person_map, with_card, without_card, fines_persons, flagged, stats


# ── Style helpers ─────────────────────────────────────────────────────────────

NAVY  = "1C2B4A"; RED   = "B91C1C"; RED_L  = "FEE2E2"
AMBER = "D97706"; AMB_L = "FEF3C7"; GREEN  = "166534"
GRN_L = "DCFCE7"; SLATE = "475569"; WHITE  = "FFFFFF"
LGREY = "F8FAFC"; BLUE_L = "EFF6FF"; BLUE  = "1D4ED8"

def _bd():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, col, val, bg=NAVY, fg=WHITE, sz=10, bold=True, align="center", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Georgia", bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _bd()

def _cell(ws, row, col, val, bg=WHITE, bold=False, fmt=None, align="left", fg="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _bd()
    if fmt:
        c.number_format = fmt

def _row_height(texts, col_widths, base=15):
    max_lines = 1
    for text, width in zip(texts, col_widths):
        if not text:
            continue
        cpl = max(1, int(width * 1.6))
        lines = max(1, -(-len(str(text)) // cpl))
        max_lines = max(max_lines, lines)
    return max(base, max_lines * 15)

def _sheet_title(ws, title_date, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = title_date
    c.font  = Font(name="Georgia", bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(person_map, with_card, without_card, fines_persons, flagged, stats, out_path):
    today_str  = date.today().strftime("%m/%d/%y")
    title_date = f"Fines & Card Verification  {today_str}"
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "⚑ Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 48
    c = ws["A1"]
    c.value = title_date.upper()
    c.font = Font(name="Georgia", bold=True, size=18, color=WHITE)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 42
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 8

    stat_boxes = [
        ("ENTRANTS",       stats["total_entrants"],              NAVY,  WHITE),
        ("FLAGGED",        stats["flagged_names"],                RED,   WHITE),
        ("SUSPENDED",      stats["susp_matches"],                 RED,   WHITE),
        ("WITH CARD #",    stats["with_card"],                    BLUE,  WHITE),
        ("WITHOUT CARD",   stats["without_card"],                 SLATE, WHITE),
        ("TOTAL $ OWED",   f"${stats['total_owed']:,.2f}",        GREEN, WHITE),
    ]
    for i, (label, val, bg, fg) in enumerate(stat_boxes, 1):
        _hdr(ws, 3, i, val,   bg=bg, fg=fg, sz=20, bold=True,  wrap=False)
        _hdr(ws, 4, i, label, bg=bg, fg=fg, sz=8,  bold=False, wrap=False)
        ws.column_dimensions[get_column_letter(i)].width = 22

    # Header row
    r = 6
    ws.row_dimensions[r].height = 22
    col_w = [24, 22, 28, 14, 14, 16]
    for ci, h in enumerate(["NAME","CLASSES ENTERED","OFFENSES/FINES",
                             "CARD #(s)","CARD EVENTS","TOTAL OWED"], 1):
        _hdr(ws, r, ci, h, bg=SLATE)
        ws.column_dimensions[get_column_letter(ci)].width = col_w[ci-1]

    # Fit to page width
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # One row per unique flagged person
    seen = set()
    for k, p in sorted(flagged.items(), key=lambda x: (x[1]["last"], x[1]["first"])):
        pk = (p["last"], p["first"])
        if pk in seen:
            continue
        seen.add(pk)

        all_c  = ", ".join(e["card_number"] for e in p["card_entries"]) or "—"
        all_ev = ", ".join(dict.fromkeys(
                     e["events"] for e in p["card_entries"] if e["events"]
                 )) or "—"
        classes   = ", ".join(sorted(p["classes"]))
        offenses  = " | ".join(e["offense"] for e in p["susp_entries"]) or "—"
        susp_total = sum(e["amount"] for e in p["susp_entries"])
        found_on  = " + ".join(filter(None, [
            "Card #"    if p["on_card_list"] else "",
            "Suspended" if p["on_susp_list"] else "",
        ]))

        bg = RED_L if p["on_susp_list"] else (BLUE_L if p["on_card_list"] else AMB_L)

        r += 1
        _cell(ws, r, 1, f"{p['last']}, {p['first']}", bg=bg, bold=True)
        _cell(ws, r, 2, classes,   bg=bg)
        _cell(ws, r, 3, offenses,  bg=bg)
        _cell(ws, r, 4, all_c,     bg=bg, align="center")
        _cell(ws, r, 5, all_ev,    bg=bg)
        if susp_total > 0:
            _cell(ws, r, 6, susp_total, bg=bg, bold=True,
                  fmt='"$"#,##0.00', align="right", fg=RED)
        else:
            _cell(ws, r, 6, "", bg=bg)
        ws.row_dimensions[r].height = _row_height(
            [f"{p['last']}, {p['first']}", classes, offenses, all_c, all_ev, ""],
            col_w)

    # Grand total
    r += 1
    _cell(ws, r, 6, f"=SUM(F7:F{r-1})", bg=GRN_L, bold=True,
          fmt='"$"#,##0.00', align="right", fg=GREEN)
    ws.merge_cells(f"A{r}:E{r}")
    gt = ws.cell(row=r, column=1, value="GRAND TOTAL")
    gt.font  = Font(name="Georgia", bold=True, size=11, color=WHITE)
    gt.fill  = PatternFill("solid", start_color=NAVY)
    gt.alignment = Alignment(horizontal="center", vertical="center")
    gt.border = _bd()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Entries With Cards
    # Every alpha-sheet name that matched a card number, A-Z
    # Red-tint rows if also suspended
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("✅ Entries With Cards")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    _sheet_title(ws2, title_date, 7)

    hdrs2 = list(zip(
        ["LAST NAME","FIRST NAME","CLASSES ENTERED",
         "CARD #(s)","CARD EVENTS","SUSPENDED?","FINES OWED"],
        [18, 16, 24, 14, 16, 14, 14]
    ))
    for ci, (h, w) in enumerate(hdrs2, 1):
        _hdr(ws2, 2, ci, h, bg=SLATE)
        ws2.column_dimensions[get_column_letter(ci)].width = w

    r2 = 2
    for k, p in sorted(with_card.items(), key=lambda x: (x[1]["last"], x[1]["first"])):
        r2 += 1
        classes   = ", ".join(sorted(p["classes"]))
        all_c     = ", ".join(e["card_number"] for e in p["card_entries"])
        all_ev    = ", ".join(dict.fromkeys(e["events"] for e in p["card_entries"] if e["events"])) or "—"
        susp_total = sum(e["amount"] for e in p["susp_entries"])
        is_susp   = p["on_susp_list"]
        bg = RED_L if is_susp else BLUE_L

        _cell(ws2, r2, 1, p["last"],    bg=bg, bold=True)
        _cell(ws2, r2, 2, p["first"],   bg=bg, bold=True)
        _cell(ws2, r2, 3, classes,      bg=bg)
        _cell(ws2, r2, 4, all_c,        bg=bg, bold=True, align="center")
        _cell(ws2, r2, 5, all_ev,       bg=bg)
        _cell(ws2, r2, 6, "YES" if is_susp else "No",
              bg=bg, bold=is_susp, align="center",
              fg=RED if is_susp else GREEN)
        if susp_total > 0:
            _cell(ws2, r2, 7, susp_total, bg=bg, bold=True,
                  fmt='"$"#,##0.00', align="right", fg=RED)
        else:
            _cell(ws2, r2, 7, "", bg=bg)
        ws2.row_dimensions[r2].height = _row_height(
            [p["last"], p["first"], classes, all_c, all_ev, "", ""],
            [w for _, w in hdrs2])

    # Total row
    r2 += 1
    ws2.merge_cells(f"A{r2}:F{r2}")
    gt2 = ws2.cell(row=r2, column=1, value=f"TOTAL WITH CARD  —  {len(with_card)} entries")
    gt2.font  = Font(name="Georgia", bold=True, size=10, color=WHITE)
    gt2.fill  = PatternFill("solid", start_color=NAVY)
    gt2.alignment = Alignment(horizontal="center", vertical="center")
    gt2.border = _bd()
    _cell(ws2, r2, 7, f"=SUM(G3:G{r2-1})", bg=GRN_L, bold=True,
          fmt='"$"#,##0.00', align="right", fg=GREEN)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Entries Without Cards
    # Every alpha-sheet name with NO card number match, A-Z
    # Red-tint if suspended
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("❌ Entries Without Cards")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A3"

    _sheet_title(ws3, title_date, 5)

    hdrs3 = list(zip(
        ["LAST NAME","FIRST NAME","CLASSES ENTERED","SUSPENDED?","FINES OWED"],
        [18, 16, 30, 14, 14]
    ))
    for ci, (h, w) in enumerate(hdrs3, 1):
        _hdr(ws3, 2, ci, h, bg=SLATE)
        ws3.column_dimensions[get_column_letter(ci)].width = w

    r3 = 2
    for k, p in sorted(without_card.items(), key=lambda x: (x[1]["last"], x[1]["first"])):
        r3 += 1
        classes   = ", ".join(sorted(p["classes"]))
        susp_total = sum(e["amount"] for e in p["susp_entries"])
        is_susp   = p["on_susp_list"]
        bg = RED_L if is_susp else LGREY

        _cell(ws3, r3, 1, p["last"],  bg=bg, bold=True)
        _cell(ws3, r3, 2, p["first"], bg=bg, bold=True)
        _cell(ws3, r3, 3, classes,    bg=bg)
        _cell(ws3, r3, 4, "YES" if is_susp else "No",
              bg=bg, bold=is_susp, align="center",
              fg=RED if is_susp else GREEN)
        if susp_total > 0:
            _cell(ws3, r3, 5, susp_total, bg=bg, bold=True,
                  fmt='"$"#,##0.00', align="right", fg=RED)
        else:
            _cell(ws3, r3, 5, "", bg=bg)
        ws3.row_dimensions[r3].height = _row_height(
            [p["last"], p["first"], classes, "", ""],
            [w for _, w in hdrs3])

    # Total row
    r3 += 1
    ws3.merge_cells(f"A{r3}:D{r3}")
    gt3 = ws3.cell(row=r3, column=1, value=f"TOTAL WITHOUT CARD  —  {len(without_card)} entries")
    gt3.font  = Font(name="Georgia", bold=True, size=10, color=WHITE)
    gt3.fill  = PatternFill("solid", start_color=NAVY)
    gt3.alignment = Alignment(horizontal="center", vertical="center")
    gt3.border = _bd()
    _cell(ws3, r3, 5, f"=SUM(E3:E{r3-1})", bg=GRN_L, bold=True,
          fmt='"$"#,##0.00', align="right", fg=GREEN)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Fine Totals
    # ALL suspended-list matches from alpha sheet
    # One row per offense, subtotal per person, grand total
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("💰 Fine Totals")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = "A3"

    _sheet_title(ws4, title_date, 7)

    hdrs4 = list(zip(
        ["LAST NAME","FIRST NAME","CLASS ENTERED",
         "OFFENSE","EVENT","FINE AMOUNT","PERSON TOTAL"],
        [16, 14, 22, 52, 10, 14, 14]
    ))
    for ci, (h, w) in enumerate(hdrs4, 1):
        _hdr(ws4, 2, ci, h, bg=SLATE)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[2].height = 30

    # Sort: highest total first
    sorted_fine_persons = sorted(
        fines_persons.items(),
        key=lambda x: (-sum(e["amount"] for e in x[1]["susp_entries"]),
                       x[1]["last"], x[1]["first"])
    )

    FINE_BG   = "FEE2E2"
    SUBTOT_BG = "475569"

    r4 = 2
    subtotal_rows = []

    for k, p in sorted_fine_persons:
        last    = p["last"]
        first   = p["first"]
        classes = ", ".join(sorted(p["classes"]))
        fines   = p["susp_entries"]

        fine_start = None
        fine_end   = None

        for fi, fine in enumerate(fines):
            r4 += 1
            if fine_start is None:
                fine_start = r4
            fine_end = r4

            show = (fi == 0)
            _cell(ws4, r4, 1, last    if show else "", bg=FINE_BG, bold=show)
            _cell(ws4, r4, 2, first   if show else "", bg=FINE_BG, bold=show)
            _cell(ws4, r4, 3, classes if show else "", bg=FINE_BG)
            _cell(ws4, r4, 4, fine["offense"], bg=FINE_BG)
            _cell(ws4, r4, 5, fine["event"],   bg=FINE_BG, align="center")
            _cell(ws4, r4, 6, fine["amount"],  bg=FINE_BG, bold=True,
                  fmt='"$"#,##0.00', align="right", fg=RED)
            _cell(ws4, r4, 7, "", bg=FINE_BG)
            ws4.row_dimensions[r4].height = _row_height(
                [last, first, classes, fine["offense"], fine["event"], "", ""],
                [w for _, w in hdrs4])

        # Subtotal row
        r4 += 1
        subtotal_rows.append(r4)
        _cell(ws4, r4, 7, f"=SUM(F{fine_start}:F{fine_end})",
              bg=GRN_L, bold=True, fmt='"$"#,##0.00', align="right", fg=GREEN)
        navy_fill = PatternFill("solid", start_color=SUBTOT_BG)
        for ci in range(1, 7):
            c = ws4.cell(row=r4, column=ci)
            c.value = None
            c.fill  = navy_fill
            c.border = _bd()
        ws4.merge_cells(f"A{r4}:F{r4}")
        sub = ws4.cell(row=r4, column=1)
        sub.value     = f"TOTAL  —  {last}, {first}"
        sub.font      = Font(name="Georgia", bold=True, size=10, color=WHITE)
        sub.fill      = navy_fill
        sub.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        sub.border    = _bd()
        ws4.row_dimensions[r4].height = 18

        # Spacer
        r4 += 1
        for ci in range(1, 8):
            ws4.cell(row=r4, column=ci).fill = PatternFill("solid", start_color="FFFFFF")
        ws4.row_dimensions[r4].height = 5

    # Grand total
    r4 += 1
    if subtotal_rows:
        formula = "+".join(f"G{sr}" for sr in subtotal_rows)
        _cell(ws4, r4, 7, f"={formula}",
              bg=GRN_L, bold=True, fmt='"$"#,##0.00', align="right", fg=GREEN)
    navy_fill2 = PatternFill("solid", start_color=SLATE)
    for ci in range(1, 7):
        c = ws4.cell(row=r4, column=ci)
        c.value = None
        c.fill  = navy_fill2
        c.border = _bd()
    ws4.merge_cells(f"A{r4}:F{r4}")
    gtr = ws4.cell(row=r4, column=1)
    gtr.value     = "GRAND TOTAL — ALL FINES"
    gtr.font      = Font(name="Georgia", bold=True, size=11, color=WHITE)
    gtr.fill      = navy_fill2
    gtr.alignment = Alignment(horizontal="center", vertical="center")
    gtr.border    = _bd()
    ws4.row_dimensions[r4].height = 22

    wb.save(out_path)
    return out_path
